import os
import sys
import time
import random
import openpyxl
import pandas as pd

from faker import Faker
from openai import OpenAI
from dotenv import load_dotenv
from openlimit import ChatRateLimiter

load_dotenv()

PROPERTIES_WORKBOOK_FILE = "./properties.xlsx"
SENTENCES_TEXT_FILE = "./sentences.txt"
GENERATED_PROPERTY_VALUES_FILE = "./generated_property_values.xlsx"
USED_GROUP_HASHES_FILE = "./used_group_hashes.txt"

MAX_NUMBER_OF_GENERATIONS = 5

GPT_MODEL = "gpt-3.5-turbo"
GPT_TEMPERATURE = 0.8
GPT_FREQUENCY_PENALTY = 0.8
GPT_TPM = 60000
GPT_RPM = 500

PROMPT_SYSTEM_MESSAGE = """
你将被提供一组属性信息，你的任务是尽你的最大可能以各种不同的形式把这组属性组装成 3 句句式多样并尽可能连贯的话。在你的回答中除了生成的句子外请不要包含任何其他的内容。
"""
PROMPT_EXAMPLE_USER_MESSAGE = """
姓名: 张三, 年龄: 20, 生日: 1990-05-04
"""
PROMPT_EXAMPLE_ASSISTANT_MESSAGE = """
1. 我的姓名是张三，今年 20 岁，出生在 1990 年 5 月 4 日，很高兴认识大家。\n\n
2. 我叫张三，今年已经 20 岁了，我出生在 1990 年 5 月 4 日。\n\n
3. 张三是我的名字，我出生在 1990 年 5 月 4 日，今年 20 岁。
"""
VALUE_PROMPT_SYSTEM_MESSAGE = """
你将被提供一种属性 (如: 日期, 班级等)，你的任务是给出 10 个该属性的示例值 (如: 1990 年 1 月 1 日, 2020 届 1 班等)。在你的回答中除了生成的示例外请不要包含任何其他的内容，每个示例值请用 "," 隔开。如果你不能给出示例值，请回答 "未知"。
"""
VALUE_PROMPT_EXAMPLE_USER_MESSAGE = """
姓名
"""
VALUE_PROMPT_EXAMPLE_ASSISTANT_MESSAGE = """
张三, 李四, 王五, 小明, 小红, 赵六, 刘七, 朱八, 郑九, 刘芳
"""

PROPERTY_GROUP_LENGTH_LOWER_BOUND = 5
PROPERTY_GROUP_LENGTH_UPPER_BOUND = 10
PROPERTY_CLASSES = {
    "address": ["地址", "住址", "地点", "位置"],
    "bank": ["网点", "银行", "分行", "支行"],
    "date_time": ["日期", "时间"],
    "phone_number": ["电话", "手机"],
}

fake = Faker("zh_CN")
fake.seed_instance(int(time.time()))
rate_limiter = ChatRateLimiter(request_limit=GPT_RPM, token_limit=GPT_TPM)
openai_client = OpenAI()

# I/O --------------------------------------------


def read_properties_workbook():
    worksheet = {}

    if not os.path.isfile(PROPERTIES_WORKBOOK_FILE):
        return worksheet

    wb = openpyxl.load_workbook(PROPERTIES_WORKBOOK_FILE)
    ws = wb.active

    for row in ws.rows:
        if row[0].value is None:
            continue
        worksheet[row[0].value] = row[1].value

    return worksheet


def read_generated_property_values_workbook():
    property_values = {}

    if not os.path.isfile(GENERATED_PROPERTY_VALUES_FILE):
        return property_values

    wb = openpyxl.load_workbook(GENERATED_PROPERTY_VALUES_FILE)

    for sheet in wb.sheetnames:
        property_values[sheet] = [cell.value for cell in wb[sheet]["A"][1:]]

    print("Previously generated property values: ", property_values)

    return property_values


def read_used_group_hashes():
    used_group_hashes = []

    if not os.path.isfile(USED_GROUP_HASHES_FILE):
        return used_group_hashes

    with open(USED_GROUP_HASHES_FILE, "r") as f:
        used_group_hashes = [int(line.strip()) for line in f.readlines()]

    print("Previously used group hashes: ", used_group_hashes)

    return used_group_hashes


def save_sentences_to_textfile(sentences):
    with open(SENTENCES_TEXT_FILE, "a") as f:
        for sentence in sentences:
            f.write(sentence + "\n")


def save_generated_property_values_to_workbook(property_values):
    print("Saving generated property values to file...")
    writer = pd.ExcelWriter(GENERATED_PROPERTY_VALUES_FILE)

    for col, values in property_values.items():
        df = pd.DataFrame({col: values})
        df.to_excel(writer, sheet_name=col, index=False)

    writer._save()


def save_used_group_hashes_to_file(used_group_hashes):
    print("Saving used group hashes to file...")
    with open(USED_GROUP_HASHES_FILE, "a") as f:
        for group_hash in used_group_hashes:
            f.write(str(group_hash) + "\n")


# Data --------------------------------------------

worksheet = read_properties_workbook()
used_group_hashes = read_used_group_hashes()
generated_property_values = read_generated_property_values_workbook()

# Logic --------------------------------------------


def generate_random_property_group(worksheet, properties):
    group = random.sample(
        properties,
        random.randint(
            PROPERTY_GROUP_LENGTH_LOWER_BOUND, PROPERTY_GROUP_LENGTH_UPPER_BOUND
        ),
    )
    group_hash = hash(tuple(sorted(group)))

    if group_hash in used_group_hashes:
        return generate_random_property_group(worksheet, properties)

    used_group_hashes.append(group_hash)

    return {prop: worksheet[prop] for prop in group}


def generate_property_group_prompt(property_group):
    prompt = ""

    for prop in property_group.items():
        val = prop[1] if prop[1] is not None else generate_random_value(prop[0])
        prompt += f"{prop[0]}: {str(val).strip()}, "

    prompt = prompt[:-2]

    return prompt


def get_openai_value_response(prop):
    openai_params = {
        "model": GPT_MODEL,
        "temperature": GPT_TEMPERATURE,
        "frequency_penalty": GPT_FREQUENCY_PENALTY,
        "messages": [
            # System message
            {"role": "system", "content": VALUE_PROMPT_SYSTEM_MESSAGE},
            # Example
            {"role": "user", "content": VALUE_PROMPT_EXAMPLE_USER_MESSAGE},
            {"role": "assistant", "content": VALUE_PROMPT_EXAMPLE_ASSISTANT_MESSAGE},
            # User message
            {"role": "user", "content": prop},
        ],
    }

    with rate_limiter.limit(**openai_params):
        response = openai_client.chat.completions.create(**openai_params)

    response_message = response.choices[0]

    if response_message.finish_reason == "stop":
        return response_message.message.content.strip()

    return None


def get_openai_response(prompt):
    openai_params = {
        "model": GPT_MODEL,
        "temperature": GPT_TEMPERATURE,
        "frequency_penalty": GPT_FREQUENCY_PENALTY,
        "messages": [
            # System message
            {"role": "system", "content": PROMPT_SYSTEM_MESSAGE},
            # Example
            {"role": "user", "content": PROMPT_EXAMPLE_USER_MESSAGE},
            {"role": "assistant", "content": PROMPT_EXAMPLE_ASSISTANT_MESSAGE},
            # User message
            {"role": "user", "content": prompt},
        ],
    }

    with rate_limiter.limit(**openai_params):
        response = openai_client.chat.completions.create(**openai_params)

    response_message = response.choices[0]

    if response_message.finish_reason == "stop":
        return response_message.message.content

    return ""


def generate_sentences(worksheet):
    print("# ---------------------------------------------------- #")
    print("# --- Pre-processing --------------------------- #")
    properties = list(worksheet.keys())
    property_group = generate_random_property_group(worksheet, properties)
    prompt = generate_property_group_prompt(property_group)

    print("# --- Main process ----------------------------- #")
    print("Generating sentences from properties: ", prompt)

    response = get_openai_response(prompt)
    sentences = get_sentences_from_openai_response(response, property_group)

    print("# --- Result ----------------------------------- #")
    print("Complete. Sentences: ", sentences)
    print("# ---------------------------------------------------- #")

    return sentences


# Helpers --------------------------------------------


def get_sentences_from_openai_response(response, property_group):
    sentences = response.split("\n")

    for idx, sentence in enumerate(sentences):
        # remove trailing and leading spaces
        sentence = sentence.strip()

        # remove leading numbers
        if sentence != "" and sentence[0].isdigit():
            sentences[idx] = sentence[3:]

    sentences = list(filter(lambda x: x != "", sentences))

    # templatize sentences
    for idx, sentence in enumerate(sentences):
        for prop in property_group.items():
            sentences[idx] = sentences[idx].replace(
                str(prop[1]).strip(), f"{{{{{prop[0]}}}}}"
            )

    return sentences


def generate_random_value(prop):
    print("Generating random value for property: ", prop)

    random_value = None

    try:
        # Read from previously generated values
        if prop in generated_property_values:
            random_value = random.choice(generated_property_values[prop])
            print("Using random value previously generated by ChatGPT: ", random_value)
            return random_value

        # Generate new values
        random_value = get_openai_value_response(prop)
        if random_value == "未知" or "抱歉" in random_value or random_value is None:
            random_value = None
        else:
            generated_property_values[prop] = random_value.strip().split(", ")
            random_value = random.choice(generated_property_values[prop])
    except:
        print("An error has occurred, using Faker...")

    if random_value is not None:
        print("Random value generated by ChatGPT: ", random_value)
        return random_value

    if any(keyword in prop for keyword in PROPERTY_CLASSES["address"]):
        random_value = fake.address()
    elif any(keyword in prop for keyword in PROPERTY_CLASSES["bank"]):
        random_value = f"{fake.company()} 银行"
    elif any(keyword in prop for keyword in PROPERTY_CLASSES["date_time"]):
        random_value = fake.date()
    elif any(keyword in prop for keyword in PROPERTY_CLASSES["phone_number"]):
        random_value = fake.phone_number()
    else:
        random_value = fake.word()

    print("Random value generated by Faker: ", random_value)
    return random_value


# Main --------------------------------------------

# Main loop
success_count = 0
while True:
    if MAX_NUMBER_OF_GENERATIONS == 0 or success_count < MAX_NUMBER_OF_GENERATIONS:
        try:
            sentences = generate_sentences(worksheet)
            save_sentences_to_textfile(sentences)
            print("")
            success_count += 1
        except KeyboardInterrupt:
            print("Keyboard interrupted, exiting...")
            sys.exit(0)
        except:
            print("Rate limit reached or an error has occurred, retrying...")
    else:
        break

# Save transactional data
if generated_property_values:
    save_generated_property_values_to_workbook(generated_property_values)

if used_group_hashes:
    save_used_group_hashes_to_file(used_group_hashes)
